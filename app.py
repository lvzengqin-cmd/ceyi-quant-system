"""
策奕量化 - 后端服务
Flask + SQLite + WebSocket实时通信 + Excel导出
"""
from flask import Flask, request, jsonify, send_from_directory, Response
from flask_cors import CORS
from flask_socketio import SocketIO, emit
import sqlite3
import hashlib
import os
import random
import threading
import time
import io
from datetime import datetime, timedelta
from functools import wraps

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

app = Flask(__name__, static_folder='../public', static_url_path='')
CORS(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# 数据库路径
DB_PATH = os.path.join(os.path.dirname(__file__), 'ceyi_quant.db')

# ============ 数据库初始化 ============
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # 用户表
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT DEFAULT 'user',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        is_active INTEGER DEFAULT 1
    )''')
    
    # 交易记录表
    c.execute('''CREATE TABLE IF NOT EXISTS trades (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT UNIQUE NOT NULL,
        user_id INTEGER,
        date TEXT NOT NULL,
        time TEXT NOT NULL,
        period TEXT NOT NULL,
        direction TEXT NOT NULL,
        entry_price REAL NOT NULL,
        exit_price REAL,
        price_diff REAL,
        amount REAL DEFAULT 100,
        breakeven_amount REAL DEFAULT 10,
        profit_rate REAL NOT NULL,
        profit REAL,
        cumulative_profit REAL,
        status TEXT DEFAULT 'pending',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        settled_at TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users (id)
    )''')
    
    # 设置表
    c.execute('''CREATE TABLE IF NOT EXISTS settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER UNIQUE,
        order_amount REAL DEFAULT 100,
        breakeven_amount REAL DEFAULT 10,
        FOREIGN KEY (user_id) REFERENCES users (id)
    )''')
    
    # 创建默认管理员
    c.execute("SELECT * FROM users WHERE username = 'admin'")
    if not c.fetchone():
        password_hash = hashlib.sha256('admin123456'.encode()).hexdigest()
        c.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)", 
                 ('admin', password_hash, 'admin'))
        conn.execute("INSERT INTO settings (user_id, order_amount, breakeven_amount) VALUES (1, 100, 10)")
    
    conn.commit()
    conn.close()
    print("✅ 数据库初始化完成")

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# ============ 认证装饰器 ============
def require_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token or not token.startswith('ceyi_token_'):
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        user_id = token.replace('ceyi_token_', '')
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE id = ? AND is_active = 1", (user_id,)).fetchone()
        conn.close()
        
        if not user:
            return jsonify({'success': False, 'error': '用户不存在或已被禁用'}), 401
        
        request.current_user = dict(user)
        return f(*args, **kwargs)
    return decorated

def require_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token or not token.startswith('ceyi_token_'):
            return jsonify({'success': False, 'error': '未授权'}), 401
        
        user_id = token.replace('ceyi_token_', '')
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE id = ? AND role = 'admin' AND is_active = 1", (user_id,)).fetchone()
        conn.close()
        
        if not user:
            return jsonify({'success': False, 'error': '需要管理员权限'}), 403
        
        request.current_user = dict(user)
        return f(*args, **kwargs)
    return decorated

# ============ 静态文件 ============
@app.route('/')
def index():
    return send_from_directory('../public', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('../public', path)

# ============ 认证API ============
@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    
    if len(username) < 3:
        return jsonify({'success': False, 'error': '用户名至少3个字符'})
    if len(password) < 6:
        return jsonify({'success': False, 'error': '密码至少6位'})
    
    conn = get_db_connection()
    existing = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        conn.close()
        return jsonify({'success': False, 'error': '用户名已存在'})
    
    password_hash = hash_password(password)
    cursor = conn.execute(
        "INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
        (username, password_hash, 'user')
    )
    conn.commit()
    
    user_id = cursor.lastrowid
    conn.execute("INSERT INTO settings (user_id, order_amount, breakeven_amount) VALUES (?, 100, 10)", (user_id,))
    conn.commit()
    conn.close()
    
    token = f"ceyi_token_{user_id}"
    return jsonify({
        'success': True,
        'token': token,
        'user': {'id': user_id, 'username': username, 'role': 'user'}
    })

@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    
    conn = get_db_connection()
    user = conn.execute(
        "SELECT * FROM users WHERE username = ? AND is_active = 1",
        (username,)
    ).fetchone()
    conn.close()
    
    if not user:
        return jsonify({'success': False, 'error': '用户不存在或已被禁用'})
    
    if hash_password(password) != user['password']:
        return jsonify({'success': False, 'error': '密码错误'})
    
    token = f"ceyi_token_{user['id']}"
    return jsonify({
        'success': True,
        'token': token,
        'user': {'id': user['id'], 'username': user['username'], 'role': user['role']}
    })

@app.route('/api/auth/profile', methods=['GET'])
@require_auth
def profile():
    user = request.current_user
    return jsonify({
        'success': True,
        'user': {'id': user['id'], 'username': user['username'], 'role': user['role']}
    })

# ============ 交易API ============
@app.route('/api/trades', methods=['GET'])
@require_auth
def get_trades():
    user = request.current_user
    page = int(request.args.get('page', 1))
    page_size = int(request.args.get('pageSize', 20))
    period = request.args.get('period', 'all')
    direction = request.args.get('direction', 'all')
    date = request.args.get('date', '')
    search = request.args.get('search', '')
    user_filter = request.args.get('userId', '')
    
    conn = get_db_connection()
    
    sql = "SELECT * FROM trades WHERE 1=1"
    params = []
    
    if user['role'] != 'admin':
        sql += " AND user_id = ?"
        params.append(user['id'])
    elif user_filter and user_filter != 'all':
        sql += " AND user_id = ?"
        params.append(user_filter)
    
    if period != 'all':
        sql += " AND period = ?"
        params.append(period)
    if direction != 'all':
        sql += " AND direction = ?"
        params.append(direction)
    if date:
        sql += " AND date = ?"
        params.append(date)
    if search:
        sql += " AND order_id LIKE ?"
        params.append(f'%{search}%')
    
    # 获取总数
    count_sql = sql.replace('SELECT *', 'SELECT COUNT(*)')
    total = conn.execute(count_sql, params).fetchone()[0]
    
    # 分页
    sql += " ORDER BY created_at DESC"
    offset = (page - 1) * page_size
    sql += f" LIMIT {page_size} OFFSET {offset}"
    
    trades = conn.execute(sql, params).fetchall()
    conn.close()
    
    return jsonify({
        'success': True,
        'data': [dict(t) for t in trades],
        'total': total,
        'page': page,
        'pageSize': page_size,
        'totalPages': (total + page_size - 1) // page_size if total > 0 else 1
    })

@app.route('/api/trades/stats', methods=['GET'])
@require_auth
def get_stats():
    user = request.current_user
    today = datetime.now().strftime('%Y-%m-%d')
    
    # 支持自定义日期范围筛选
    date = request.args.get('date', '')
    start_date = request.args.get('startDate', '')
    end_date = request.args.get('endDate', '')
    
    conn = get_db_connection()
    
    base_sql = "SELECT * FROM trades WHERE 1=1"
    params = []
    if user['role'] != 'admin':
        base_sql += " AND user_id = ?"
        params.append(user['id'])
    
    # 根据筛选条件确定统计范围
    if date:
        # 单日统计
        filter_sql = base_sql + " AND date = ?"
        filter_params = params + [date]
    elif start_date and end_date:
        # 区间统计
        filter_sql = base_sql + " AND date >= ? AND date <= ?"
        filter_params = params + [start_date, end_date]
    else:
        # 今日统计（默认）
        filter_sql = base_sql + " AND date = ?"
        filter_params = params + [today]
    
    filtered_trades = conn.execute(filter_sql, filter_params).fetchall()
    
    # 基础统计
    total_orders = len(filtered_trades)
    long_count = len([t for t in filtered_trades if t['direction'] == 'long'])
    short_count = len([t for t in filtered_trades if t['direction'] == 'short'])
    
    # 已结算订单统计（用于胜率计算）
    settled_sql = base_sql + " AND status != 'pending'"
    settled_params = params.copy()
    
    if date:
        settled_sql += " AND date = ?"
        settled_params.append(date)
    elif start_date and end_date:
        settled_sql += " AND date >= ? AND date <= ?"
        settled_params.extend([start_date, end_date])
    
    settled = conn.execute(settled_sql, settled_params).fetchall()
    win_count = len([t for t in settled if t['status'] == 'win'])
    lose_count = len([t for t in settled if t['status'] == 'lose'])
    win_rate = (win_count / len(settled) * 100) if settled else 0
    
    # 利润统计（使用筛选范围内的已结算订单）
    profit_sum = sum([t['profit'] or 0 for t in settled if t['profit'] is not None])
    
    # 累计利润（全部已结算订单）
    all_settled = conn.execute(
        base_sql.replace("SELECT *", "SELECT *") + " AND status != 'pending'",
        params
    ).fetchall()
    total_profit = sum([t['profit'] or 0 for t in all_settled if t['profit'] is not None])
    
    conn.close()
    
    return jsonify({
        'success': True,
        'data': {
            'todayOrders': total_orders,
            'todayLongCount': long_count,
            'todayShortCount': short_count,
            'winRate': f'{win_rate:.1f}%',
            'winCount': win_count,
            'loseCount': lose_count,
            'settledCount': len(settled),
            'todayProfit': round(profit_sum, 2),
            'totalProfit': round(total_profit, 2),
            'filterRange': {
                'date': date,
                'startDate': start_date,
                'endDate': end_date
            }
        }
    })

@app.route('/api/trades/export', methods=['GET'])
@require_auth
def export_trades():
    """导出交易记录为CSV或Excel"""
    export_format = request.args.get('format', 'csv')
    period = request.args.get('period', 'all')
    direction = request.args.get('direction', 'all')
    date = request.args.get('date', '')
    start_date = request.args.get('startDate', '')
    end_date = request.args.get('endDate', '')
    
    conn = get_db_connection()
    
    sql = "SELECT * FROM trades WHERE 1=1"
    params = []
    if user['role'] != 'admin':
        sql += " AND user_id = ?"
        params.append(user['id'])
    if period != 'all':
        sql += " AND period = ?"
        params.append(period)
    if direction != 'all':
        sql += " AND direction = ?"
        params.append(direction)
    if date:
        sql += " AND date = ?"
        params.append(date)
    if start_date:
        sql += " AND date >= ?"
        params.append(start_date)
    if end_date:
        sql += " AND date <= ?"
        params.append(end_date)
    
    sql += " ORDER BY created_at DESC"
    trades = conn.execute(sql, params).fetchall()
    conn.close()
    
    if export_format == 'xlsx' and EXCEL_AVAILABLE:
        return export_to_excel(trades)
    else:
        return export_to_csv(trades)

def export_to_excel(trades):
    """导出为Excel格式"""
    wb = Workbook()
    ws = wb.active
    ws.title = "交易记录"
    
    # 样式定义
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['序号', '日期', '订单ID', '时间', '周期', '方向', '入场价', '结算价', '价差', 
               '投入金额', '利润率', '利润', '累计利润', '结果']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 数据行
    for row_idx, t in enumerate(trades, 2):
        ws.cell(row=row_idx, column=1, value=row_idx-1).border = border
        ws.cell(row=row_idx, column=2, value=t['date']).border = border
        ws.cell(row=row_idx, column=3, value=t['order_id']).border = border
        ws.cell(row=row_idx, column=4, value=t['time']).border = border
        ws.cell(row=row_idx, column=5, value=t['period']).border = border
        
        dir_cell = ws.cell(row=row_idx, column=6, value='做多' if t['direction']=='long' else '做空')
        dir_cell.border = border
        
        ws.cell(row=row_idx, column=7, value=t['entry_price']).border = border
        ws.cell(row=row_idx, column=8, value=t['exit_price'] or '').border = border
        ws.cell(row=row_idx, column=9, value=t['price_diff'] or '').border = border
        ws.cell(row=row_idx, column=10, value=t['amount']).border = border
        ws.cell(row=row_idx, column=11, value=f"{t['profit_rate']*100:.0f}%").border = border
        ws.cell(row=row_idx, column=12, value=t['profit'] if t['profit'] is not None else '').border = border
        ws.cell(row=row_idx, column=13, value=t['cumulative_profit'] if t['cumulative_profit'] is not None else '').border = border
        
        status_map = {'win': '盈利', 'lose': '亏损', 'pending': '待结算'}
        status_cell = ws.cell(row=row_idx, column=14, value=status_map.get(t['status'], t['status']))
        status_cell.border = border
        
        # 利润为负时红色显示
        if t['profit'] is not None and t['profit'] < 0:
            ws.cell(row=row_idx, column=12).font = Font(color="EF4444")
    
    # 设置列宽
    col_widths = [8, 12, 18, 10, 8, 8, 12, 12, 10, 10, 10, 12, 12, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64+i) if i <= 26 else 'A'+chr(64+i-26)].width = width
    
    # 保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'策奕量化交易记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename*=UTF-8\'\'{filename}'}
    )

def export_to_csv(trades):
    """导出为CSV格式"""
    csv = "序号,日期,订单ID,时间,周期,方向,入场价,结算价,价差,投入金额,利润率,利润,累计利润,结果\n"
    for i, t in enumerate(trades):
        csv += f'{i+1},{t["date"]},{t["order_id"]},{t["time"]},{t["period"]},'
        csv += f'{"做多" if t["direction"]=="long" else "做空"},'
        csv += f'{t["entry_price"]},{t["exit_price"] or ""},{t["price_diff"] or ""},'
        csv += f'{t["amount"]},{t["profit_rate"]*100:.0f}%,'
        csv += f'{t["profit"] or ""},{t["cumulative_profit"] or ""},'
        csv += f'{"盈利" if t["status"]=="win" else ("亏损" if t["status"]=="lose" else "待结算")}\n'
    
    filename = f'策奕量化交易记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    return Response(
        csv,
        mimetype='text/csv; charset=utf-8-sig',
        headers={'Content-Disposition': f'attachment; filename*=UTF-8\'\'{filename}'}
    )

# ============ Webhook API ============
@app.route('/api/webhook/<period>/<direction>', methods=['POST', 'GET'])
def webhook(period, direction):
    """接收交易信号"""
    if period not in ['5m', '10m', '30m']:
        return jsonify({'success': False, 'error': '无效周期'})
    if direction not in ['long', 'short']:
        return jsonify({'success': False, 'error': '无效方向'})
    
    # 获取BTC价格 (模拟)
    btc_price = 65000 + random.uniform(-500, 500)
    
    conn = get_db_connection()
    
    # 获取今日序号
    today = datetime.now().strftime('%Y-%m-%d')
    seq = conn.execute("SELECT COUNT(*) FROM trades WHERE date = ?", (today,)).fetchone()[0] + 1
    
    order_id = f"CEYI{today.replace('-', '')}{seq:03d}"
    
    # 利润率配置
    profit_rates = {'5m': 0.50, '10m': 0.75, '30m': 1.00}
    
    # 获取admin用户
    admin = conn.execute("SELECT id FROM users WHERE role = 'admin'").fetchone()
    user_id = admin['id'] if admin else 1
    
    now = datetime.now()
    time_str = now.strftime('%H:%M:%S')
    
    # 插入交易
    conn.execute('''INSERT INTO trades 
        (order_id, user_id, date, time, period, direction, entry_price, profit_rate, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
        (order_id, user_id, today, time_str, period, direction, btc_price, profit_rates[period], 'pending'))
    conn.commit()
    trade_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.close()
    
    # 广播新订单
    socketio.emit('new_trade', {
        'id': trade_id,
        'order_id': order_id,
        'period': period,
        'direction': direction,
        'entry_price': btc_price
    })
    
    # 延迟结算
    delay_seconds = int(period.replace('m', '')) * 60
    threading.Thread(target=delayed_settle, args=(trade_id, delay_seconds)).start()
    
    return jsonify({
        'success': True,
        'message': f'订单 {order_id} 已创建，{delay_seconds}秒后结算',
        'order_id': order_id,
        'entry_price': btc_price
    })

def delayed_settle(trade_id, delay):
    time.sleep(delay)
    settle_trade(trade_id)

def settle_trade(trade_id):
    """结算交易"""
    conn = get_db_connection()
    
    trade = conn.execute("SELECT * FROM trades WHERE id = ?", (trade_id,)).fetchone()
    if not trade or trade['status'] != 'pending':
        conn.close()
        return
    
    # 模拟价格变动
    price_change = (random.random() - 0.5) * 1500
    exit_price = trade['entry_price'] + price_change
    
    price_diff = exit_price - trade['entry_price']
    actual_diff = price_diff if trade['direction'] == 'long' else -price_diff
    profit = trade['amount'] * trade['profit_rate'] * (actual_diff / trade['entry_price'])
    
    status = 'win' if profit >= 0 else 'lose'
    
    conn.execute('''UPDATE trades SET 
        exit_price = ?, price_diff = ?, profit = ?, status = ?, settled_at = ?
        WHERE id = ?''',
        (exit_price, price_diff, profit, status, datetime.now().isoformat(), trade_id))
    
    # 重新计算累计利润
    settled_trades = conn.execute(
        "SELECT * FROM trades WHERE status != 'pending' ORDER BY created_at"
    ).fetchall()
    
    cumulative = 0
    for t in settled_trades:
        cumulative += t['profit'] or 0
        conn.execute("UPDATE trades SET cumulative_profit = ? WHERE id = ?", (cumulative, t['id']))
    
    conn.commit()
    conn.close()
    
    socketio.emit('trade_settled', {
        'id': trade_id,
        'order_id': trade['order_id'],
        'profit': profit,
        'status': status
    })

# ============ 设置API ============
@app.route('/api/settings', methods=['GET'])
@require_auth
def get_settings():
    user = request.current_user
    conn = get_db_connection()
    settings = conn.execute("SELECT * FROM settings WHERE user_id = ?", (user['id'],)).fetchone()
    conn.close()
    
    return jsonify({
        'success': True,
        'data': dict(settings) if settings else {'order_amount': 100, 'breakeven_amount': 10}
    })

@app.route('/api/settings', methods=['PUT'])
@require_auth
def update_settings():
    user = request.current_user
    data = request.json
    
    conn = get_db_connection()
    conn.execute('''INSERT OR REPLACE INTO settings (user_id, order_amount, breakeven_amount)
        VALUES (?, ?, ?)''',
        (user['id'], data.get('order_amount', 100), data.get('breakeven_amount', 10)))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# ============ 用户管理API ============
@app.route('/api/admin/stats', methods=['GET'])
@require_admin
def get_admin_stats():
    """获取管理员统计数据"""
    conn = get_db_connection()
    
    total_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    today_new = conn.execute("SELECT COUNT(*) FROM users WHERE DATE(created_at) = DATE('now', 'localtime')").fetchone()[0]
    active_users = conn.execute("SELECT COUNT(DISTINCT user_id) FROM trades").fetchone()[0]
    total_trades = conn.execute("SELECT COUNT(*) FROM trades").fetchone()[0]
    
    conn.close()
    
    return jsonify({
        'success': True,
        'data': {
            'totalUsers': total_users,
            'activeUsers': active_users,
            'totalTrades': total_trades,
            'todayNewUsers': today_new
        }
    })

@app.route('/api/admin/users', methods=['GET'])
@require_admin
def get_users():
    conn = get_db_connection()
    users = conn.execute("""
        SELECT u.id, u.username, u.role, u.created_at, u.is_active,
               COUNT(t.id) as trade_count,
               COALESCE(SUM(CASE WHEN t.status = 'win' THEN t.profit ELSE 0 END), 0) as total_profit
        FROM users u
        LEFT JOIN trades t ON u.id = t.user_id
        GROUP BY u.id
        ORDER BY u.created_at DESC
    """).fetchall()
    conn.close()
    
    return jsonify({'success': True, 'data': [dict(u) for u in users]})

@app.route('/api/admin/users/<int:user_id>/toggle', methods=['POST'])
@require_admin
def toggle_user(user_id):
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    
    if user and user['username'] != 'admin':
        new_status = 0 if user['is_active'] else 1
        conn.execute("UPDATE users SET is_active = ? WHERE id = ?", (new_status, user_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'is_active': new_status})
    
    conn.close()
    return jsonify({'success': False, 'error': '无法修改管理员'})

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@require_admin
def delete_user(user_id):
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    
    if user and user['username'] != 'admin':
        conn.execute("DELETE FROM trades WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM settings WHERE user_id = ?", (user_id,))
        conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    
    conn.close()
    return jsonify({'success': False, 'error': '无法删除管理员'})

# ============ 健康检查 ============
@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'service': '策奕量化', 'time': datetime.now().isoformat()})

# ============ Socket.IO ============
@socketio.on('connect')
def on_connect():
    print('🔌 客户端连接')

@socketio.on('disconnect')
def on_disconnect():
    print('🔌 客户端断开')

# ============ 初始化 ============
init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"🚀 策奕量化服务器启动中，端口: {port}")
    socketio.run(app, host='0.0.0.0', port=port, debug=False, allow_unsafe_werkzeug=True)